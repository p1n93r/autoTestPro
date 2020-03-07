from utils import HTMLTestRunner
import openpyxl
import os,time


def printScreen(driver,fileName):
    '''
    Function:利用selenium的驱动截图
    '''
    filePath=getProjectRootPath()+"report/image/"+fileName+".png"
    driver.get_screenshot_as_file(filePath)


def getProjectRootPath():
    '''
    Function:获取项目的根路径
    Notice:此写法中，有硬编码，如果想移植，需要手动修改此函数中的硬编码：autoTestPro(项目名称)
    '''
    #为了保证通用性，将windows下的文件路径改写成linux下的
    currentPath=os.path.abspath('.').replace("\\","/")
    #得到项目根路径
    rootPaht=currentPath.split("/autoTestPro")[0]+"/autoTestPro/"
    return rootPaht
   
   
def createHtmlReport(reportTitle,testSuite):
    '''
    Function:执行测试用例，生成基于html的报告
    '''
    nowTime=time.strftime("%Y%m%d_%H_%M_%S")
    fileName=getProjectRootPath()+"report/html/"+nowTime+".html"
    fp=open(fileName,'wb')
    HTMLTestRunner.HTMLTestRunner(stream=fp,title=reportTitle,description='测试环境：OS:windows.  Browser:Firefox.  Framework:selenium.  Language:python').run(testSuite)
    fp.close()
    
    
def readExcel(filePath):
    '''
    Function:读取excel文件，返回一个二维表
    '''
    wb = openpyxl.load_workbook(filePath)
    #获取当前活跃的表单
    ws = wb.active
    #获取总行数和总列数
    rows=ws.max_row
    columns=ws.max_column
    #定义待返回的列表
    result=[]
    for i in range(1,rows+1):
        rowValue=[]
        for j in range(1,columns+1):
            rowValue.append(ws.cell(row=i, column=j).value)
        result.append(rowValue)
    return result
    
def writeExcel(filePath,*info):
    '''
    Function:改写特定excel的特定单元格内容
    Notice:info=(row,col,cellValue),其中row和col从0开始
    '''
    wb = openpyxl.load_workbook(filePath)
    #获取当前活跃的表单
    ws = wb.active
    ws.cell(row=info[0]+1, column=info[1]+1).value=info[2]
    #保存
    wb.save(filePath)
    

if __name__=="__main__":
    createHtmlReport()